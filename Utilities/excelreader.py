import openpyxl


def get_excel_data(path, sheet_name):
    workbook = openpyxl.load_workbook(filename=path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column

   # print("total row:" + str(total_rows) + "------total column: " + str(total_columns) )
    data_table = []
    for row in range(2, total_rows+1):
        item = []
        for column in range(1, total_columns + 1):
            item.append(str(sheet.cell(row, column).value))

        data_table.append(item)

  #  print(data_table)

    return data_table


get_excel_data("..//Excel//data.xlsx", "Registration")